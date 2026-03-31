from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    'Company', 'Job Title', 'Location', 'Description',
    'Status', 'Applied Date', 'Last Updated', 'Thread IDs',
]

# Excel column indices (1-based)
COL = {h: i for i, h in enumerate(HEADERS, 1)}

STATUS_COLORS = {
    'Applied':             'BDD7EE',  # blue
    'Interview Scheduled': 'FFEB9C',  # yellow
    'Offer Received':      'C6EFCE',  # green
    'Rejected':            'FFC7CE',  # red/pink
    'Withdrawn':           'D9D9D9',  # grey
}

COLUMN_WIDTHS = [22, 32, 22, 55, 22, 14, 14, 0]  # 0 = hidden


class ExcelManager:
    def __init__(self, filepath='job_applications.xlsx'):
        self.filepath = Path(filepath)
        self.wb = self._load_or_create()
        self.ws = self.wb.active

    # ------------------------------------------------------------------
    # Setup
    # ------------------------------------------------------------------

    def _load_or_create(self):
        if self.filepath.exists():
            return openpyxl.load_workbook(self.filepath)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Job Applications'
        self._write_headers(ws)
        wb.save(self.filepath)
        return wb

    def _write_headers(self, ws):
        header_fill = PatternFill(fill_type='solid', fgColor='243F60')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for col, name in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col, value=name)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'

        for i, width in enumerate(COLUMN_WIDTHS, 1):
            letter = get_column_letter(i)
            if width == 0:
                ws.column_dimensions[letter].hidden = True
            else:
                ws.column_dimensions[letter].width = width

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def get_all_jobs(self):
        jobs = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                jobs.append({
                    'company':      row[COL['Company'] - 1],
                    'job_title':    row[COL['Job Title'] - 1],
                    'location':     row[COL['Location'] - 1],
                    'description':  row[COL['Description'] - 1],
                    'status':       row[COL['Status'] - 1],
                    'applied_date': row[COL['Applied Date'] - 1],
                    'last_updated': row[COL['Last Updated'] - 1],
                    'thread_ids':   row[COL['Thread IDs'] - 1] or '',
                })
        return jobs

    def has_thread_id(self, thread_id):
        for row in self.ws.iter_rows(
            min_row=2,
            min_col=COL['Thread IDs'],
            max_col=COL['Thread IDs'],
            values_only=True,
        ):
            if thread_id in str(row[0] or ''):
                return True
        return False

    def add_job(self, company, job_title, location, description,
                status, applied_date, thread_id=''):
        today = datetime.now().strftime('%Y-%m-%d')
        # Insert a blank row just below the header so newest entries appear at the top.
        self.ws.insert_rows(2)
        row_num = 2
        values = [
            company, job_title, location, description,
            status, applied_date or today, today, thread_id,
        ]
        for col, value in enumerate(values, 1):
            cell = self.ws.cell(row=row_num, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        self._color_row(row_num, status)
        self.wb.save(self.filepath)
        print(f'  + Added   : {company} — {job_title} [{status}]')

    def update_status(self, row_index, new_status, thread_id=''):
        """row_index is 0-based (position in get_all_jobs() list)."""
        actual_row = row_index + 2
        today = datetime.now().strftime('%Y-%m-%d')

        self.ws.cell(row=actual_row, column=COL['Status'], value=new_status)
        self.ws.cell(row=actual_row, column=COL['Last Updated'], value=today)

        if thread_id:
            tid_cell = self.ws.cell(row=actual_row, column=COL['Thread IDs'])
            existing = str(tid_cell.value or '')
            if thread_id not in existing:
                tid_cell.value = (existing + ',' + thread_id).strip(',')

        self._color_row(actual_row, new_status)
        self.wb.save(self.filepath)

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _color_row(self, row_num, status):
        color = STATUS_COLORS.get(status, 'FFFFFF')
        fill = PatternFill(fill_type='solid', fgColor=color)
        for col in range(1, len(HEADERS) + 1):
            self.ws.cell(row=row_num, column=col).fill = fill
